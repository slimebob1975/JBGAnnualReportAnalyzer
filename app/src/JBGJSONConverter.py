import csv
import json
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.src import JBGMetricSchema as schema
from app.src import JBGValidation as validation
from app.src.JBGAnnualReportAnalysis import JBGAnnualReportAnalyzer
from app.src.JBGFundNames import FundNameResolver

logger = logging.getLogger(__name__)

class JsonConverter:
    def __init__(self, json_path: str | Path, include_sources: bool = False):
        self.json_path = Path(json_path)
        if not self.json_path.exists():
            raise FileNotFoundError(f"JSON file not found: {self.json_path}")
        self.include_sources = include_sources
        self.data = self._load_json()

    METADATA_PREFIX = "_"

    def _load_json(self):
        with open(self.json_path, encoding='utf-8') as f:
            return json.load(f)

    def _funds(self) -> dict:
        """The fund entries only. Keys prefixed with an underscore hold
        metadata such as the validation findings, not a fund."""
        return {
            name: years
            for name, years in self.data.items()
            if not name.startswith(self.METADATA_PREFIX) and isinstance(years, dict)
        }

    def findings(self) -> list[dict]:
        """Validation findings recorded in the result file, if any."""
        recorded = self.data.get("_rimlighetskontroller") or []
        return recorded if isinstance(recorded, list) else []

    def _findings_by_cell(self) -> dict:
        index = {}
        for finding in self.findings():
            for metric in finding.get("berörda_nyckeltal") or []:
                key = (finding.get("kassa"), str(finding.get("år")), metric)
                index.setdefault(key, []).append(finding.get("kontroll", ""))
        return index

    def _rows(self) -> list[dict]:
        """Flatten to Fund | Year | Key | Value [| Source | Certainty | Comment].

        The certainty and comment the model produces used to be discarded here,
        even though the prompt spends considerable effort calibrating them.
        """
        rows = []
        flagged = self._findings_by_cell()
        for fund_name, years in self._funds().items():
            for year, key_numbers in years.items():
                for key, value_dict in key_numbers.items():
                    if not isinstance(value_dict, dict):
                        continue
                    row = {
                        "Fund": fund_name,
                        "Year": year,
                        "Key": key,
                        "Value": value_dict.get(JBGAnnualReportAnalyzer.FIELD_VALUE),
                    }
                    if self.include_sources:
                        row["Source"] = value_dict.get(JBGAnnualReportAnalyzer.FIELD_SOURCE)
                        row["Certainty"] = value_dict.get(
                            JBGAnnualReportAnalyzer.FIELD_CERTAINTY
                        )
                        row["Comment"] = value_dict.get(
                            JBGAnnualReportAnalyzer.FIELD_COMMENT
                        )
                        row["Validering"] = "; ".join(
                            flagged.get((fund_name, str(year), key), [])
                        )
                    rows.append(row)
        return rows

    @property
    def columns(self) -> list[str]:
        base = ["Fund", "Year", "Key", "Value"]
        if not self.include_sources:
            return base
        return base + ["Source", "Certainty", "Comment", "Validering"]

    def to_dataframe(self):
        """Optional convenience wrapper. Requires the 'analysis' extra.

        pandas is imported here rather than at module level: it was a hard
        dependency of the whole package purely so that to_csv could write a
        semicolon-separated file, which the standard library does fine.
        """
        try:
            import pandas as pd
        except ImportError as ex:  # pragma: no cover
            raise ImportError(
                "to_dataframe() kräver pandas. Installera med: pip install '.[analysis]'"
            ) from ex
        return pd.DataFrame(self._rows(), columns=self.columns)

    def to_csv(self, output_path: str | Path):
        output_path = Path(output_path)
        rows = self._rows()
        # utf-8-sig so Excel on Windows opens it with the right encoding, and
        # semicolons because that is what a Swedish locale expects.
        with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(
                handle, fieldnames=self.columns, delimiter=";", extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(rows)
        logger.info(f"CSV file saved to {output_path} ({len(rows)} rader)")

    def to_excel(self, output_path: str | Path, by: str = "fund"):
        """
        Save to Excel with multiple sheets.
        by: 'fund' or 'year'
        """
        import pandas as pd  # only needed for this optional flat export

        df = self.to_dataframe()
        output_path = Path(output_path)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            if by == "fund":
                for fund, group in df.groupby("Fund"):
                    group.to_excel(writer, sheet_name=self._sanitize_sheetname(fund), index=False)
            elif by == "year":
                for year, group in df.groupby("Year"):
                    group.to_excel(writer, sheet_name=str(year), index=False)
            else:
                raise ValueError("Parameter 'by' must be either 'fund' or 'year'")

        logger.info(f"Excel with sheets by '{by}' saved to {output_path}")

    # Confidence bands used to shade value cells. Read straight off the
    # "säkerhet" the model reports, which the prompt calibrates explicitly.
    CERTAINTY_BANDS = [
        (schema.CERTAINTY_EXPLICIT, "C6EFCE", "explicit – står ordagrant i dokumentet"),
        (schema.CERTAINTY_DERIVED, "FFEB9C", "härledd – uträknad eller tolkad rubrik"),
        (schema.CERTAINTY_UNCERTAIN, "FFC7CE", "osäker – bör kontrolleras mot källan"),
    ]
    FLAGGED_FILL = "E1BEE7"

    @classmethod
    def _certainty_fill(cls, certainty) -> PatternFill | None:
        """Shade by the reported level, mapping legacy floats onto the scale."""
        level = schema.certainty_level(certainty)
        for name, colour, _ in cls.CERTAINTY_BANDS:
            if level == name:
                return PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        return None

    def to_excel_by_year(
        self,
        output_path: str | Path,
        key_def_path: str | Path,
        fund_names: None | str | Path = None,
        findings: list | None = None,
    ):
        """
        Export JSON data to Excel with:
        - One sheet per year
        - Funds as columns
        - Nyckeltal as rows, grouped and ordered by key_def_path
        - Value cells shaded by the model's own reported certainty
        - The source, certainty and comment attached as a cell note
        - A separate sheet listing failed sanity checks, if any
        """
        with open(key_def_path, encoding="utf-8") as f:
            key_defs = json.load(f)

        # Resolve fund names through the shared resolver rather than an exact
        # dict lookup, which matched none of the names in a real sample.
        resolver = None
        if fund_names:
            try:
                resolver = FundNameResolver(fund_names)
            except (OSError, json.JSONDecodeError) as ex:
                logger.warning(f"Kunde inte läsa kassaregistret {fund_names}: {ex}")

        def display_name(fund: str) -> str:
            return resolver.short_name(fund) if resolver else fund

        grouped_keys = {}
        for entry in key_defs:
            group = entry.get("Grupp", "🧩 Övrigt")
            grouped_keys.setdefault(group, []).append(entry["Nyckeltal"])
        group_order = list(grouped_keys.keys())
        all_keys = [entry["Nyckeltal"] for entry in key_defs]

        if findings:
            flagged = validation.findings_by_cell(findings)
        else:
            # Fall back to whatever the result file recorded, so exporting an
            # existing JSON keeps the flags.
            flagged = {
                key: [
                    validation.Finding(
                        fund=key[0], year=key[1], rule=rule, message="", metrics=[key[2]]
                    )
                    for rule in rules
                ]
                for key, rules in self._findings_by_cell().items()
            }

        # Build year -> fund -> key -> entry dict (not just the bare value, so
        # the source, certainty and comment survive to this point)
        year_structured = {}
        for fund, year_data in self._funds().items():
            for year, metrics in year_data.items():
                per_fund = year_structured.setdefault(str(year), {}).setdefault(fund, {})
                for key in all_keys:
                    entry = metrics.get(key)
                    per_fund[key] = entry if isinstance(entry, dict) else None

        wb = Workbook()
        del wb["Sheet"]

        for year in sorted(year_structured):
            fund_data = year_structured[year]
            ws = wb.create_sheet(title=str(year))
            funds = sorted(fund_data.keys(), key=display_name)

            header = ["Nyckeltal"]
            for fund in funds:
                header.append(display_name(fund))
                if self.include_sources:
                    header.append("källa")
            ws.append(header)
            for col_num in range(1, len(header) + 1):
                ws.cell(row=1, column=col_num).font = Font(bold=True)
            ws.freeze_panes = "B2"

            row_idx = 2
            for group in group_order:
                ws.cell(row=row_idx, column=1, value=group).font = Font(bold=True)
                row_idx += 1

                for key in grouped_keys[group]:
                    ws.cell(row=row_idx, column=1, value=key)
                    col_idx = 2
                    for fund in funds:
                        entry = fund_data.get(fund, {}).get(key) or {}
                        value = entry.get(JBGAnnualReportAnalyzer.FIELD_VALUE)
                        certainty = entry.get(JBGAnnualReportAnalyzer.FIELD_CERTAINTY)
                        source = entry.get(JBGAnnualReportAnalyzer.FIELD_SOURCE, "")
                        comment = entry.get(JBGAnnualReportAnalyzer.FIELD_COMMENT, "")

                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        problems = flagged.get((fund, str(year), key), [])
                        if problems:
                            cell.fill = PatternFill(
                                start_color=self.FLAGGED_FILL,
                                end_color=self.FLAGGED_FILL,
                                fill_type="solid",
                            )
                        else:
                            fill = self._certainty_fill(certainty)
                            if fill is not None:
                                cell.fill = fill

                        note = self._build_note(certainty, source, comment, problems)
                        if note:
                            # Cell notes keep the model's reasoning available on
                            # hover without adding three columns per fund.
                            cell.comment = Comment(note, "JBG nyckeltalsanalys")
                            cell.comment.width = 380
                            cell.comment.height = 180
                        col_idx += 1

                        if self.include_sources:
                            ws.cell(row=row_idx, column=col_idx, value=source)
                            col_idx += 1
                    row_idx += 1

            self._autosize(ws)

        self._write_legend_sheet(wb, findings or [], resolver)
        wb.save(output_path)
        logger.info(f"Excel file saved to {output_path}")

    @staticmethod
    def _build_note(certainty, source, comment, problems) -> str:
        parts = []
        if certainty:
            parts.append(f"Säkerhet: {certainty}")
        if source:
            parts.append(f"Källa: {source}")
        if comment:
            parts.append(f"Kommentar: {comment}")
        for problem in problems:
            parts.append(f"⚠ {problem.rule}: {problem.message}")
        return "\n\n".join(parts)

    @staticmethod
    def _autosize(ws) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            longest = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=0,
            )
            ws.column_dimensions[letter].width = max(8, min(longest + 2, 40))

    def _write_legend_sheet(self, wb, findings: list, resolver) -> None:
        """A short reading guide, plus any sanity checks that failed."""
        ws = wb.create_sheet(title="Läsanvisning")
        ws.append(["Färgkodning av värden (modellens egen bedömning)"])
        ws["A1"].font = Font(bold=True)
        for _, colour, label in self.CERTAINTY_BANDS:
            ws.append([label])
            ws.cell(row=ws.max_row, column=1).fill = PatternFill(
                start_color=colour, end_color=colour, fill_type="solid"
            )
        ws.append(["Ingår i en misslyckad rimlighetskontroll"])
        ws.cell(row=ws.max_row, column=1).fill = PatternFill(
            start_color=self.FLAGGED_FILL, end_color=self.FLAGGED_FILL, fill_type="solid"
        )
        ws.append([])
        ws.append(["Håll pekaren över ett värde för källa, säkerhet och kommentar."])
        ws.append([])

        ws.append(["Rimlighetskontroller"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        if not findings:
            ws.append(["Inga anmärkningar."])
        else:
            ws.append(["Kassa", "År", "Kontroll", "Anmärkning"])
            for col in range(1, 5):
                ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
            for finding in findings:
                ws.append([
                    resolver.short_name(finding.fund) if resolver else finding.fund,
                    finding.year,
                    finding.rule,
                    finding.message,
                ])
        self._autosize(ws)

    def _sanitize_sheetname(self, name: str) -> str:
        # Excel sheet names max 31 chars and cannot contain some symbols
        return name[:31].replace("/", "-").replace("\\", "-").replace(":", "-")


